#!/usr/bin/env python3
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data = json.load(open('data/menu_normalizado.json', encoding='utf-8'))
rows = data['filas']

GOLD = 'D9BB73'      # brand primary
DARK = '04111D'      # brand dark
LIGHT = 'F5F0E6'

hdr_font = Font(bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill('solid', fgColor=DARK)
cat_font = Font(bold=True, color=DARK, size=12)
cat_fill = PatternFill('solid', fgColor=GOLD)
sub_font = Font(bold=True, color=DARK, size=10, italic=True)
sub_fill = PatternFill('solid', fgColor='EDE3C7')
wrap = Alignment(wrap_text=True, vertical='top')
thin = Side(style='thin', color='DDDDDD')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ---------- RESUMEN ----------
ws = wb.active
ws.title = 'RESUMEN'
ws.sheet_view.showGridLines = False
title = ws.cell(1, 1, 'PANISSE — Extracción completa del menú')
title.font = Font(bold=True, size=18, color=DARK)
ws.cell(2, 1, 'Fuente: menupp.co  ·  Sede: Mall Pilares del Bosque, Pereira  ·  Moneda: COP').font = Font(size=10, italic=True, color='666666')

menus_order = []
for r in rows:
    if r['menu'] not in menus_order:
        menus_order.append(r['menu'])

hrow = 4
for i, h in enumerate(['Menú', 'Categorías', 'Productos', 'Con imagen', 'Nota'], 1):
    c = ws.cell(hrow, i, h); c.font = hdr_font; c.fill = hdr_fill; c.border = border
notes = {
    'BRUNCH': 'Menú principal (catálogo completo)',
    'LUNCH Y DINNER': 'Casi idéntico a BRUNCH (247 productos compartidos)',
    'VINOS': 'Mostrado como "VINOS Y BEBIDAS" en la web',
    'DOMICILIO': 'Menú de domicilio (inactivo en la web)',
}
rr = hrow + 1
for m in menus_order:
    mrows = [x for x in rows if x['menu'] == m]
    ws.cell(rr, 1, m).border = border
    ws.cell(rr, 2, len({x['categoria_ruta'] for x in mrows})).border = border
    ws.cell(rr, 3, len(mrows)).border = border
    ws.cell(rr, 4, sum(1 for x in mrows if x['num_imagenes'] > 0)).border = border
    ws.cell(rr, 5, notes.get(m, '')).border = border
    rr += 1
uniq = {x['product_id'] for x in rows}
ws.cell(rr + 1, 1, 'Productos ÚNICOS (sin duplicar entre menús):').font = Font(bold=True, color=DARK)
ws.cell(rr + 1, 3, len(uniq)).font = Font(bold=True, color=DARK)
ws.cell(rr + 2, 1, 'Total de filas (productos × menús):').font = Font(color='666666')
ws.cell(rr + 2, 3, len(rows)).font = Font(color='666666')
ws.cell(rr + 4, 1, 'Cada hoja de menú lista los productos agrupados por categoría, en el mismo orden que la web.').font = Font(italic=True, size=9, color='888888')
ws.cell(rr + 5, 1, 'Las imágenes están descargadas en la carpeta images/ (nombre en la columna "Imagen").').font = Font(italic=True, size=9, color='888888')
for col, w in {'A': 34, 'B': 12, 'C': 12, 'D': 12, 'E': 52}.items():
    ws.column_dimensions[col].width = w

# ---------- PER-MENU SHEETS ----------
headers = ['Categoría', 'Subcategoría', 'Producto', 'Descripción', 'Precio', 'Variantes de precio', 'Recom.', 'Nuevo', 'Imagen (archivo)']
keys = ['categoria_raiz', 'subcategoria', 'producto', 'descripcion', 'precio', 'variantes_precio', 'recomendado', 'nuevo', 'imagen_archivo']
widths = [24, 22, 30, 60, 12, 30, 8, 7, 34]

def add_menu_sheet(name, mrows):
    safe = name[:31]
    ws = wb.create_sheet(safe)
    ws.sheet_view.showGridLines = False
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h); c.font = hdr_font; c.fill = hdr_fill; c.border = border; c.alignment = Alignment(vertical='center')
    ws.freeze_panes = 'A2'
    r = 2
    last_cat = last_sub = None
    for x in mrows:
        # category band
        if x['categoria_raiz'] != last_cat:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
            c = ws.cell(r, 1, f"▸ {x['categoria_raiz']}"); c.font = cat_font; c.fill = cat_fill
            r += 1; last_cat = x['categoria_raiz']; last_sub = None
        if x['subcategoria'] and x['subcategoria'] != last_sub:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
            c = ws.cell(r, 1, f"    · {x['subcategoria']}"); c.font = sub_font; c.fill = sub_fill
            r += 1; last_sub = x['subcategoria']
        for i, k in enumerate(keys, 1):
            v = x[k]
            if k == 'recomendado' or k == 'nuevo':
                v = '★' if v == 'sí' else ''
            c = ws.cell(r, i, v); c.alignment = wrap; c.border = border
            if k == 'precio':
                c.font = Font(bold=True, color=DARK)
        r += 1
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

for m in menus_order:
    add_menu_sheet(m, [x for x in rows if x['menu'] == m])

# ---------- UNIQUE PRODUCTS SHEET ----------
seen = {}
for x in rows:
    if x['product_id'] not in seen:
        seen[x['product_id']] = dict(x, menus=set())
    seen[x['product_id']]['menus'].add(x['menu'])
uheaders = ['Menús', 'Categoría', 'Subcategoría', 'Producto', 'Descripción', 'Precio', 'Variantes', 'Imagen (archivo)']
ukeys = ['menus', 'categoria_raiz', 'subcategoria', 'producto', 'descripcion', 'precio', 'variantes_precio', 'imagen_archivo']
ws = wb.create_sheet('PRODUCTOS ÚNICOS')
ws.sheet_view.showGridLines = False
for i, h in enumerate(uheaders, 1):
    c = ws.cell(1, i, h); c.font = hdr_font; c.fill = hdr_fill; c.border = border
ws.freeze_panes = 'A2'
r = 2
for x in sorted(seen.values(), key=lambda z: (z['categoria_raiz'], z['subcategoria'], z['producto'])):
    x['menus'] = ', '.join(sorted(x['menus']))
    for i, k in enumerate(ukeys, 1):
        c = ws.cell(r, i, x[k]); c.alignment = wrap; c.border = border
    r += 1
for i, w in enumerate([24, 22, 20, 30, 55, 12, 28, 34], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save('data/MENU_PANISSE.xlsx')
print('Wrote data/MENU_PANISSE.xlsx with sheets:', [ws.title for ws in wb.worksheets])
